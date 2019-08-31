# -*- coding:utf-8 -*-
# @Author: guoyq
# @Last Modified time: 2019-08-30
# 科技讯官网：http://www.kejixun.com/


# 爬虫相关模块
import pymysql
import requests
from bs4 import BeautifulSoup
# 文件相关模块
import os
from openpyxl import load_workbook
# 邮件相关模块
import smtplib
from email.header import Header
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
# 微信相关模块
import itchat
# 时间相关模块
import time
import datetime
import schedule
'''
程序说明：
1、该程序作用为从新闻网站爬取新闻资讯，每天将资讯信息定时发送到需要的邮箱、微信朋友和微信群，并将每天的新闻信息
   存储入数据库，以供后续可能的数据分析模块调用。同时建立了日志信息，便于自动运行过程中出现问题时排查原因。
2、要发送的邮箱，微信朋友和微信群从excel表格中获取(有模板)，方便随时增减需要发送的人员和群，避免直接改动程序。
3、附属必要文件有:(1)、news.sql(建立数据库)；(2)、新闻收发基础信息.xlsx(发送人信息)；(3)、红心.png(作为邮件附件)。
4、被次程序是爬取的科技讯新闻网站的新闻资讯，通过稍微改动爬虫类（class Spider）很容易换成其他新闻网站的资讯……
5、为保证程序每天自动运行，所使用的登录邮箱和密码是直接写死的，第一次需要配置（只能程序中更改，虽然可以在excel中
   或其他配置文件中读取，但由于密码信息直接存放在excel表格中太过明显，容易泄漏，不好），同时也避免每次输入密码的
   麻烦，不使用input输入。
'''


class Tim:
    # 时间类，主要用于获取当前时间
    def __init__(self):
        self.now_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def get_news_time(self):   # 打印新闻获取时间
        print('新闻获取时间：%s' % self.now_time)
        return self.now_time

    @staticmethod
    def get_day():   # 获取到天
        today = datetime.datetime.now().strftime('%mx%dy').replace('x', '月').replace('y', '日')
        return today


class LogPath:
    # 文件和路径类，用于创建日志目录和日志文件
    @staticmethod
    def mk_main_dir():   # 创建主日志目录, 查看基础路径是否存在，不存在则创建
        path = r'D:\news_log'
        if os.path.exists(path):
            print('%s 路径已存在' % path)
        else:
            os.makedirs(path)
            print('%s 创建成功' % path)

    @staticmethod
    def mk_log_dir():   # 检查分日志目录，不存在则创建
        path_list = [
            r'D:\news_log\mysql',
            r'D:\news_log\email',
            r'D:\news_log\wechat',
            r'D:\news_log\spider',
            r'D:\news_log\other'
        ]
        for path in path_list:
            if os.path.exists(path):
                print('%s 路径已存在' % path)
            else:
                os.makedirs(path)
                print('%s 创建成功' % path)

    @staticmethod
    def log_file(path, name, e):   # 创建日志文件，在此处设置模板，方便其他模块直接调用生成日志文件
        err_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
        file_time = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        with open(r'D:\news_log\%s\%s%s.txt' % (path, file_time, name), 'a', encoding='utf-8') as f:
            f.writelines('\n%s  %s' % (err_time, e))


class Mysql:
    # mysql数据库类，部分方法用于添加数据备用，主要通过insert_news()方法将每日新闻写入数据库
    def __init__(self):   # 连接数据库，获取游标(光标)
        self.db = pymysql.connect(host='localhost', port=3306, user='root', password='123456', db='news', charset='utf8')
        self.cursor = self.db.cursor()

    def close_cursor(self):   # 关闭游标方法
        return self.cursor.close()

    def close_db(self):   # 关闭数据库连接方法
        return self.db.close()

    def execute_sql(self, sql):   # 平时不用，用于临时想要执行sql语句时使用
        try:
            self.cursor.execute(sql)
            self.db.commit()
        except Exception as e:
            print('sql语句执行失败！', e)
            self.db.rollback()
            LogPath().log_file('mysql', 'execute', e)

    def insert_user(self):   # 平时不用，用于临时插入user表
        sql = 'insert into user(user_id,user_name,user_psd,sex,age,hobby) values(%s,%s,%s,%s,%s,%s)'
        data = (0, 'guoyq', '123456', '男', '30', '旅游，骑行')
        try:
            self.cursor.execute(sql, data)
            self.db.commit()
        except Exception as e:
            print('user数据插入失败！', e)
            self.db.rollback()
            LogPath().log_file('mysql', 'sql_user', e)

    def insert_email(self):  # 平时不用，用于临时插入email基础信息表
        sql = 'insert into email(email_id,email_host,email_port,login_email,login_psd,send_email,send_info) values(%s,%s,%s,%s,%s,%s,%s)'
        data = (0, 'smtp.163.com', 465, 'abc@163.com', '*****', 'xyz@163.com', '科技讯今日早报')
        try:
            self.cursor.execute(sql, data)
            self.db.commit()
        except Exception as e:
            print('email数据插入失败！', e)
            self.db.rollback()
            LogPath().log_file('mysql', 'sql_email', e)

    def insert_wechat(self):   # 平时不用，用于临时插入wechat基础信息表
        sql = 'insert into wechat(wechat_id,login_user,send_user,send_room,send_info) values(%s,%s,%s,%s,%s)'
        data = (0, '你的微信昵称', '朋友的微信昵称', '微信群名称', '科技讯今日早报')
        try:
            self.cursor.execute(sql, data)
            self.db.commit()
        except Exception as e:
            print('wechat数据插入失败！', e)
            self.db.rollback()
            LogPath().log_file('mysql', 'sql_wechat', e)

    def insert_news(self, news_info, news_day, news_time):   # 主要方法，用于将爬虫爬取的新闻信息插入数据库
        sql = 'insert into kejixun(kejx_id,today,title,brief,news_url,from_url,send_time) values(%s,%s,%s,%s,%s,%s,%s)'
        data = (0, news_day, news_info[1], news_info[2], news_info[3], 'http://www.kejixun.com/', news_time)
        try:
            self.cursor.execute(sql, data)
            self.db.commit()
        except Exception as e:
            print('news数据插入失败！', e)
            self.db.rollback()
            LogPath().log_file('mysql', 'sql_news', e)


class Spider:
    # 爬虫类，主要用于获取科技讯官网新闻资讯
    def __init__(self):  # 列出需要请求的链接和设置请求头
        self.headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.103 Safari/537.36'}
        self.url = 'http://www.kejixun.com/'

    def post_url(self):  # 获取网页
        res = requests.get(self.url, headers=self.headers)
        print('首页请求状态：%s' % res.status_code)
        # res.encoding='utf-8'   # 根据需要限定编码，确保后续的解析能得到正确结果
        return res

    def parser(self):  # 解析网页
        res = self.post_url()
        html = BeautifulSoup(res.text, 'html.parser')
        return html

    def get_navs(self):  # 解析以获取首页导航栏,备用
        html = self.parser()
        nav_list = []
        navs = html.find('div', class_='main-tit clearfix').find_all('li')
        for i in navs:
            nav = i.find('a').text
            nav_list.append(nav)
        return nav_list

    def get_titles(self):  # 解析以获取首页最新新闻标题，内容简介和链接信息
        html = self.parser()
        # 以下三个列表分别用于存放爬取的不同内容，分别供邮件，微信和数据库调用
        title_list = []
        title_list_nourl = []
        news_info_list = []
        news_num = 1
        print('科技讯最新新闻资讯如下：')
        try:
            titles = html.find('div', id='content').find_all('li')
            for j in titles:   # 获取最新资讯的标题，内容简介和内容的url地址
                title = j.find('div', class_='detail-info').find('a').text
                brief = j.find('div', class_='detail-info').find('p').text
                new_url = j.find('div', class_='detail-info').find('a')['href']
                # 新闻默认有100条数据，我们只要他的前10条，太多懒得看
                if news_num <= 10:
                    print('%s、%s\n%s\n%s' % (news_num, title, brief, new_url))
                    title_list.append('%s、%s\n%s\n%s\n' % (news_num, title, brief, new_url))
                    title_list_nourl.append('%s、%s\n' % (news_num, title))
                    news_info_list.append((news_num, title, brief, new_url))
                else:
                    pass
                news_num += 1
        except Exception as e:
            print('get_titles数据解析失败！%s' % e)
            LogPath().log_file('spider', 'get_titles', e)
        return title_list, title_list_nourl, news_info_list


class Email:
    # 邮件类，主要用于将爬取的信息邮件发送到需要的邮箱中
    def __init__(self, title_list):   # 注意：my_email是用于发送的邮箱，my_psd一般是邮箱的授权码，更改发送邮箱时配对更改
        self.host = 'smtp.163.com'    # 163邮箱的域名，其他邮箱需要查询对应的域名信息
        self.port = 465               # 163邮箱的端口号
        self.my_email = 'abcdefg@163.com'   # 登录邮箱，下面的授权码需配对才能发出邮件，可以百度搜索如何获取xx邮箱授权码
        self.my_psd = 'y%se4%s2' % ('12345', '1234567')   # 授权码（不是密码），此处为避免太过明显，简单复杂化，你也可以直接赋值授权码
        self.title_list = title_list

    @staticmethod
    def get_account():   # 获取excel表格中的邮件sheet页的信息
        try:
            wb = load_workbook(r'.\文件\新闻收发基础信息.xlsx')
            sh_email = wb.get_sheet_by_name('邮件')
            # 获取B3-B202的数据信息
            all_info = sh_email['B3:B202']
            email_list = []
            for content in all_info:
                email = content[0].value
                if email is not None:
                    email_list.append(email)
            # 将email信息封装为邮件可接受的情况
            to_email = ','.join(email_list)
            wb.close()
        except Exception as e:
            print('请检查文件是否正常', e)
            LogPath().log_file('other', 'mail_account', e)
            to_email = ''
        return to_email

    def send_info(self):  # 邮件发送的信息
        now_tims = Tim()
        msg = MIMEMultipart()
        # 邮件正文，将爬取的新闻标题放在正文中
        content = '\n'.join(self.title_list)
        text = MIMEText(content, 'plain', 'utf-8')
        # 发送一张压轴图片
        image_name = '红心.png'
        path = r'.\文件\%s' % image_name
        image = MIMEApplication(open(path, 'rb').read())
        image.add_header('Content-Disposition', 'attachment', filename=image_name)
        # 传送数据
        msg.attach(image)
        msg.attach(text)
        # 邮件标题信息
        msg['from'] = Header(self.my_email)
        msg['to'] = Header(self.get_account())
        msg['subject'] = Header('科技讯今日早报%s' % now_tims.get_day())
        return msg

    def send_mail(self):   # 发送邮件
        server = smtplib.SMTP_SSL()
        server.connect(self.host, self.port)
        print('连接成功！')
        server.login(self.my_email, self.my_psd)
        print('登录成功！')
        try:
            print('稍候，正在准备发送邮件！')
            server.sendmail(self.my_email, self.get_account().split(','), self.send_info().as_string())
            print('邮件发送成功^-^')
            server.quit()
        except Exception as e:
            print(e)
            print('邮件发送失败，请检查！')
            server.quit()
            LogPath().log_file('email', 'send_mail', e)


class WeChat:
    # 微信类,参考网址：https://itchat.readthedocs.io/zh/latest/
    def __init__(self, title_list_nourl):
        # 初始化发送人和群
        self.title_list_nourl = title_list_nourl
        self.today = Tim().get_day()

    @staticmethod
    def get_account():   # 从excel表格中获取登录的微信昵称，朋友昵称和朋友圈名称
        try:
            wb = load_workbook(r'.\文件\新闻收发基础信息.xlsx')
            # 获取邮件表格
            sh_wechat = wb.get_sheet_by_name('微信')
            # 获取B3-B202的数据信息
            all_friend_info = sh_wechat['C3:C202']
            all_room_info = sh_wechat['D3:D202']
            # 存储从excel中读取的微信朋友昵称和微信群名称
            friend_list = []
            room_list = []
            # 获取微信朋友昵称
            for content1 in all_friend_info:
                friend = content1[0].value
                if friend is not None:
                    friend_list.append(friend)
            # 获取微信群昵称
            for content2 in all_room_info:
                room = content2[0].value
                if room is not None:
                    room_list.append(room)
            wb.close()
        except Exception as e:
            print('请检查文件是否正常', e)
            LogPath().log_file('other', 'wechat_account', e)
            friend_list = ''
            room_list = ''
        return friend_list, room_list

    @staticmethod
    def auto_login():   # 自动登陆
        print('准备登陆微信')
        itchat.auto_login(hotReload=True, enableCmdQR=False)
        print('微信登陆成功')

    def send_friend_info(self):   # 发送信息给朋友
        print('开始发送信息到个人')
        for friend in self.get_account()[0]:
            user_name = itchat.search_friends(name=friend)[0]['UserName']
            content1 = '科技讯今日早报  %s\n' % self.today
            content2 = '--【http://www.kejixun.com】\n'
            content3 = ''.join(self.title_list_nourl)
            itchat.send_msg(content1+content2+content3, user_name)

    def send_room_info(self):   # 发送信息给微信群
        print('开始发送信息到微信群')
        for room in self.get_account()[1]:
            room_name = itchat.search_chatrooms(name=room)[0]['UserName']
            content4 = '科技讯今日早报  %s\n' % self.today
            content5 = '--【http://www.kejixun.com】\n'
            content6 = ''.join(self.title_list_nourl)
            itchat.send_msg(content4+content5+content6, room_name)

    @staticmethod
    def keep_login():   # 保持登录状态
        print('保持微信登陆')
        itchat.auto_login(hotReload=True)


def run_all():
    # 整体执行函数
    # 时间函数部分
    print('开始')
    times = Tim()
    # 日志目录和文件部分
    log_path = LogPath()
    log_path.mk_main_dir()
    log_path.mk_log_dir()
    # 时间函数
    news_time = times.get_news_time()
    news_day = times.get_day()
    # 爬取新闻标题信息部分
    spider = Spider()
    title_list = spider.get_titles()
    # 将新闻信息插入数据库
    mysql = Mysql()
    try:
        num_ins = 1
        for news_info in title_list[2]:
            mysql.insert_news(news_info, news_day, news_time)
            print('向数据库插入第%s条数据！' % num_ins)
            num_ins += 1
    except Exception as e:
        print('插入数据库报错信息:%s' % e)
        LogPath().log_file('mysql', 'sql_all', e)
    finally:
        mysql.close_cursor()
        mysql.close_db()
    # 邮件发送部分
    try:
        emails = Email(title_list[0])
        emails.send_mail()
    except Exception as e:
        print('写入邮件报错信息:%s' % e)
        LogPath().log_file('email', 'send_mail', e)
    # 微信发送部分
    try:
        wechat = WeChat(title_list[1])
        wechat.auto_login()
        wechat.send_friend_info()
        wechat.send_room_info()
        wechat.keep_login()
    except Exception as e:
        print('写入微信报错信息:%s' % e)
        LogPath().log_file('wechat', 'wechat', e)
    print('完成')


if __name__ == "__main__":
    # 主函数,引用schedule定时器，定时执行
    print('开始执行')
    schedule.every().day.at("08:10").do(run_all)
    while True:
        schedule.run_pending()
        time.sleep(1)
