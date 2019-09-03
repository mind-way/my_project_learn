# -*- coding:utf-8 -*-
# @Author: guoyq
# @Last Modified time: 2019-07-28
# 安居客成都所有新楼盘信息爬取+存储+结巴分词+词云
# 安居客成都楼盘网址：https://cd.fang.anjuke.com/loupan/all/
import gevent
from gevent import monkey
monkey.patch_all()
import requests
import openpyxl
import jieba
import jieba.analyse
import matplotlib.pyplot as plt
import numpy as np
import time
import random
from bs4 import BeautifulSoup
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from wordcloud import WordCloud, ImageColorGenerator
from PIL import Image as Img   # from PIL import Image的Image或和 from openpyxl.drawing.image import Image的Image冲突！
from gevent.queue import Queue


'''
一、本程序作用为爬取安居客网站中成都市所有的新楼盘信息，并对获取的信息进行处理，得出成都市各楼盘的分布情况，
    通过词云的方式直观显示，同时将楼盘基本信息和对用楼盘宣传照片存入excel，以便后续做其他分析；
二、涉及技术主要有：1、代理IP获取、验证并组成IP代理池(通过爬虫获取)；2、爬虫、反反爬与多线程(requests+BeautifulSoup+gevent)
    3、excel存取(openpyxl)，如果有必要，也可以存储在mysql或者MongoDB数据库中；4、结巴分词(jieba)；5、numpy图片处理(numpy)；
    6、云词及背景加载(wordcloud+matplotlib+PIL)；以上主要分三大部分写：代理IP部分+主爬虫部分+数据处理部分
三、为了绕开安居客信息各种反爬手段，本程序从网上公开的代理网站获取代理IP，组成代理IP池，随机选取IP进行动态爬取，避免被封IP。
    程序的异常信息分爬虫分析，结巴分词和词云三部分存储为txt日志。
'''


# 一、代理IP获取、验证和选取
def get_ip_list(url):
    # 获取代理IP，可以通过https://www.xicidaili.com/wt/ 获取ip_list，也可以通过其他代理网站获取，代码稍作修改即可
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.139 Safari/537.36'
        }
    res = requests.get(url, headers=headers)
    res = BeautifulSoup(res.text, 'html.parser')
    results = res.select('#ip_list tr')
    for result in results[1:]:
        ip = result.select('td')[1].text
        port = result.select('td')[2].text
        judge(ip, port)


ip_list = []


def judge(ip, port):
    # 判断获取的IP是否有效
    global ip_list
    proxy = {'http':ip+':'+port}
    try:
        res = requests.get('https://www.baidu.com', proxies=proxy)
    except Exception:
        print('该ip：'+ip+'无效')
        return False
    else:
        if 200 <= res.status_code < 300:
            ip_list.append((ip, port))
            return True
        else:
            print('该ip：' + ip + '无效')
            return False


def get_random_ip():
    # 从ip_list中随机获取一个ip，因为代理IP有存活周期，所有此处不用将所有可用的代理IP存起来，即用即取即可，存起来也可以。
    ip, port = random.choice(ip_list)
    result = judge(ip, port)
    if result:
        return ip + ':' + port
    else:
        ip_list.remove((ip, port))


# 二、主爬虫部分，安居客信息爬取和存储
def crawler():
    # 主爬虫函数,获取安居客成都所有新楼盘基本信息
    N = 1
    # 从队列中取出url,并进行解析
    while not work.empty():
        url = work.get_nowait()
        header = work_h.get_nowait()
        res = requests.get(url, headers=header, proxies=proxy)
        print('请求状态%s' % res.status_code)
        if 200<= res.status_code <300:
            html = BeautifulSoup(res.text, 'html.parser')
            # 提取数据
            all_infos = html.find('div', class_='key-list imglazyload').find_all('div', class_='item-mod')
            for infos in all_infos:
                try:
                    build_name = infos.find('div', class_='infos').find('span', class_='items-name').text
                    room_price = infos.find('a', class_='favor-pos').find('p').text
                    build_addr = infos.find('div', class_='infos').find('a', class_='address').find('span', class_='list-map').text
                    build_type = infos.find('div', class_='infos').find('div', class_='tag-panel').find('i', class_='status-icon wuyetp').text
                    sale_state = infos.find('div', class_='infos').find('div', class_='tag-panel').find('i').text
                    build_state =infos.find('div', class_='infos').find('div', class_='tag-panel').find('span').text
                    room_area = infos.find('div', class_='infos').find('a', class_='huxing').text.strip().replace(' ', '')
                    picture = infos.find('a', class_='pic').find('img')['src']
                    link = infos.find('div', class_='infos').find('a', class_='lp-name')['href']
                    # 输出数据
                    print('%s,楼盘名称:%s，房子价格:%s，地址:%s，建筑类型:%s，销售状态:%s，建筑状态:%s，房子面积:%s，楼盘图片:%s，楼盘链接:%s'
                          %(N, build_name, room_price, build_addr, build_type, sale_state, build_state, room_area, picture, link))
                    # 如果需要在excel中写入图片再下载写入，否则不下载图片
                    if wr_pic == 'y':
                        # 图片文件转化为二进制文件下载
                        try:
                            myheader = {
                                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.139 Safari/537.36'
                            }
                            res_picture = requests.get(picture, headers=myheader, proxies=proxy)
                            pic = res_picture.content  # 图片的二进制数据
                            # 下载图片
                            photo = open(r'E:\练习文件\图片\安居客楼盘信息\%s' % (picture[-53:].replace('/', '-')), 'wb')
                            photo.write(pic)
                            # 获取pic的二进制内容
                            photo.close()
                            # 准备写入图片文件
                            time.sleep(9)
                            img = Image(r'E:\练习文件\图片\安居客楼盘信息\%s' % (picture[-53:].replace('/', '-')))
                            # p = sheet.add_image(img)
                        except Exception as e:
                            print(e)
                            print('图片下载或写入excel可能出错，请检查！')
                            img = Image(r'E:\练习文件\图片\安居客楼盘信息\1.jpg')
                            err_time0 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                            with open(r'E:\练习文件\项目文件\python_log\spider_analysis.txt', 'a', encoding='utf-8') as f_ana1:
                                f_ana1.writelines('图片下载或写入excel报错：%s  %s\n' % (e, err_time0))
                        # 楼盘信息写入excel
                        try:
                            rows = [build_name, room_price, build_addr, build_type, sale_state, build_state, room_area, picture, link, sheet1.add_image(img, 'J%s'%(N+1))]
                            sheet1.append(rows)
                            wb.save(r'E:\练习文件\安居客成都所有新楼盘信息.xlsx')
                            N += 1
                        except Exception as e:
                            print(e)
                            print('安居客信息写入Excel可能出错，请检查！')
                            err_time1 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                            with open(r'E:\练习文件\项目文件\python_log\spider_analysis.txt', 'a', encoding='utf-8') as f_ana2:
                                f_ana2.writelines('安居客信息写入excel报错：%s  %s\n' % (e, err_time1))
                    else:
                        try:
                            rows = [build_name, room_price, build_addr, build_type, sale_state, build_state, room_area, picture, link, '/']
                            sheet1.append(rows)
                            wb.save(r'E:\练习文件\安居客成都所有新楼盘信息.xlsx')
                            N += 1
                        except Exception as e:
                            print(e)
                            print('安居客信息写入Excel可能出错，请检查！')
                            err_time2 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                            with open(r'E:\练习文件\项目文件\python_log\spider_analysis.txt', 'a', encoding='utf-8') as f_ana3:
                                f_ana3.writelines('安居客信息写入excel报错：%s  %s\n' % (e, err_time2))
                except Exception as e:
                    print(e)
                    print('解析可能出错，请检查！')
                    err_time3 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                    with open(r'E:\练习文件\项目文件\python_log\spider_analysis.txt', 'a', encoding='utf-8') as f_ana4:
                        f_ana4.writelines('url解析失败报错：%s  %s\n' % (e, err_time3))
        else:
            print('向安居客网站请求失败！')


# 三、excel预处理+结巴分词+词云显示
# 3.1 数据预处理
def read_xlsx(path_xlsx):
    #  读取excel文件
    wb = load_workbook(path_xlsx)
    wb.guess_types = True   # 猜测格式类型
    sheet2 = wb.get_sheet_by_name('安居客新楼盘信息')
    return sheet2


def write_txt(path_txt, sheet):
    # 文件需要处理部分存储为txt文档，也可以直接赋值给变量，此处这么写的目的是避免大文件直接赋值到变量（可调整读取量）
    with open(path_txt, 'w', encoding='utf-8') as f:
        for column in sheet['C']:         # 读取指定列
            f.writelines(str(column.value)+'\n')
            # print(column.value)


def read_txt(path_txt):
    # 从txt中提取文本信息
    with open(path_txt, 'r', encoding='utf-8') as f2:
        content = f2.read()
    return content


# 3.2 结巴分词部分
def userdict(path_userdict):
    # 自定义词典
    jieba.load_userdict(path_userdict)


def set_stop_words(path_stop_words):
    # 设置停用词
    jieba.analyse.set_stop_words(path_stop_words)


def jieba_cut(content):
    # 精确模式
    jb_cut = jieba.cut(content, cut_all=False)
    return '/'.join(jb_cut)


def write_extract(path_txt, content):
    # 提取关键词并保存
    with open(path_txt, 'w', encoding='utf-8') as f3:
        tags = jieba.analyse.extract_tags(content, topK=100, withWeight=True)
        keywords = {}
        for v, n in tags:
            keywords[v] = n
            f3.writelines('%s:%s' % (v, n)+'\n')
            print('%s:%s' % (v, n))
    return keywords


# 3.3 背景图处理
def np_img(path_img):
    # 提供背景图,通过numpy处理
    bg_img = np.array(Img.open(path_img))
    return bg_img


# 3.4 词云处理和显示
def my_wordcloud(path_font, bg_img, keywords):
    # 词云处理和显示
    wordcloud = WordCloud(
        # 添加字体路径         # msyh.ttc(微软雅黑)/msyhl.ttc(微软雅黑-细)/msyhbd.ttc(微软雅黑-细)/simhei.ttf(黑体)
        font_path=path_font,   # simsun.ttc(宋体)/stxihei.ttf(华文细黑)/SIMLI.ttf(隶书)/stcaiyun.ttf(华文彩云)
        # 设置背景色，高宽
        background_color='white', mask=bg_img, max_font_size=130, min_font_size=4).generate_from_frequencies(keywords)
    image_colors = ImageColorGenerator(bg_img)
    plt.imshow(wordcloud.recolor(color_func=image_colors), interpolation='bilinear')
    plt.axis('off')
    plt.show()


if __name__ == "__main__":
    answer = input('是先爬取最新数据进行分析还是用老数据分析，任意键爬取最新数据，老数据分析输入n:')
    if answer != 'n':
        # 实例化线程，准备进行多线程操作
        work = Queue()
        work_h = Queue()

        # 一、获取代理IP
        print('====================1、开始获取代理IP……====================')
        get_ip_list('https://www.xicidaili.com/wt/')

        # 二、获取url地址列表和请求头列表
        print('===================2、准备url地址和请求列表……====================')
        url_list = []
        h_list = []
        for x in range(35, 37):
            # 构造url地址
            p_url = 'https://cd.fang.anjuke.com/loupan/all/p%s/' % x  # https://cd.fang.anjuke.com/loupan/all/p3/
            url_list.append(p_url)
            # 构造请求头信息，请求头中referer参数在变化，此处根据变化规律一同构造出来
            headers = {
                'referer': 'https://cd.fang.anjuke.com/loupan/all/p%s/' % (x - 1),
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.103 Safari/537.36'
            }
            h_list.append(headers)
            # 代理IP地址信息
            proxy = {
                'http': get_random_ip()
            }
        print('url列表如下：%s' % url_list)
        print('请求头列表如下：%s' % h_list)

        # 三、写入excel标题栏信息，详细信息在爬虫函数中写入
        print('===================3、写入excel标题栏信息====================')
        wb = openpyxl.Workbook()
        sheet1 = wb.active
        sheet1.title = '安居客新楼盘信息'   # 也可以通过
        titles = ['楼盘名称', '房子价格', '地址', '建筑类型', '销售状态', '建筑状态', '房子面积', '楼盘图片', '楼盘链接', '楼盘图片']
        sheet1.append(titles)

        # 四、将url地址列表加入队列,将请求头加入队列，添加任务并执行
        for real_url in url_list:
            work.put_nowait(real_url)
        for head in h_list:
            work_h.put_nowait(head)
        # 添加任务
        task_list = []
        for i in range(1):
            task = gevent.spawn(crawler)
            task_list.append(task)
        # 执行任务
        startspider = input('是否开始执行爬取安居客信息？是请输入y,任意键退出:')
        if startspider == 'y':
            print('===================4、开始爬取安居客成都新楼盘信息====================')
            wr_pic = input('excel中是否要写入图片信息？任意键不写入，需要输入y：')
            gevent.joinall(task_list)
            # 执行完毕后可以在此处开始云词等方面内容处理（推荐），也可以不在此处处理，利用以前的数据做处理。
        else:
            wb.close()
        wb.close()
    else:
        pass

    # 五、excel数据预处理
    startcloud = input('是否开始执行结巴云词处理？是请输入y,任意键退出：')
    if startcloud == 'y':
        print('====================5、开始预处理数据====================')
        path_xlsx = r'E:\练习文件\安居客成都所有新楼盘信息.xlsx'
        path_txt = r'E:\练习文件\安居客成都所有新楼盘信息.txt'
        # 读取的excel信息
        try:
            sheet = read_xlsx(path_xlsx)
        except Exception as e:
            print(e)
            print('检查文件是否被打开，关闭文件重新处理')
            err_time4 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
            with open(r'E:\练习文件\项目文件\python_log\jieba_analysis.txt', 'a', encoding='utf-8') as f_jieba1:
                f_jieba1.writelines('读取excel信息进行预处理时报错：%s  %s\n' % (e, err_time4))
            time.sleep(20)
            sheet = read_xlsx(path_xlsx)
        # 需要的部分写入txt
        try:
            write_txt(path_txt, sheet)
        except Exception as e:
            print(e)
            print('检查文件是否被打开，关闭文件重新处理')
            err_time5 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
            with open(r'E:\练习文件\项目文件\python_log\jieba_analysis.txt', 'a', encoding='utf-8') as f_jieba2:
                f_jieba2.writelines('写入txt预处理信息报错：%s  %s\n' % (e, err_time5))
            time.sleep(20)
            write_txt(path_txt, sheet)
        # 读取txt信息准备给结巴进行处理
        try:
            content = read_txt(path_txt)
        except Exception as e:
            print(e)
            print('检查文件是否被打开，关闭文件重新处理')
            err_time6 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
            with open(r'E:\练习文件\项目文件\python_log\jieba_analysis.txt', 'a', encoding='utf-8') as f_jieba3:
                f_jieba3.writelines('读取写入的txt预处理信息报错：%s  %s\n' % (e, err_time6))
            time.sleep(20)
            content = read_txt(path_txt)

        # 六、结巴分词
        print('====================6、开始结巴分词====================')
        path_userdict = r'E:\练习文件\项目文件\anjudict.txt'
        path_stop_words = r'E:\练习文件\项目文件\anju-stop.txt'
        # 加载用户字典和停用词，内容需要根据结巴实际可能没有添加的专有名词和无用词语进行添加
        try:
            userdict(path_userdict)
            set_stop_words(path_stop_words)
        except Exception as e:
            print(e)
            print('检查文件是否被打开，关闭文件重新处理')
            err_time7 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
            with open(r'E:\练习文件\项目文件\python_log\jieba_analysis.txt', 'a', encoding='utf-8') as f_jieba4:
                f_jieba4.writelines('用户字典和停用词加载报错：%s  %s\n' % (e, err_time7))
            time.sleep(20)
            userdict(path_userdict)
            set_stop_words(path_stop_words)

        # 结巴分词
        try:
            cut = jieba_cut(content)
            print(cut)
            keywords = write_extract(path_txt, content)
        except Exception as e:
            print('检查文件路径或文件是否被打开，关闭文件重新处理')
            err_time8 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
            with open(r'E:\练习文件\项目文件\python_log\jieba_analysis.txt', 'a', encoding='utf-8') as f_jieba5:
                f_jieba5.writelines('结巴分词报错：%s  %s\n' % (e, err_time8))
            time.sleep(20)
            cut = jieba_cut(content)
            print(cut)
            keywords = write_extract(path_txt, content)

        # 七、numpy对背景图进行处理
        print('===================7、开始处理词云需要的背景图信息====================')
        try:
            path_img = r'E:\练习文件\图片\词云用图\cd1.jpg'
            bg_img = np_img(path_img)
        except Exception as e:
            print(e)
            print('检查图片路径是否错误！')
            err_time9 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
            with open(r'E:\练习文件\项目文件\python_log\cloud_analysis.txt', 'a', encoding='utf-8') as f_cloud1:
                f_cloud1.writelines('numpy背景图处理报错：%s  %s\n' % (e, err_time9))
            path_img = r'E:\练习文件\图片\词云用图\1.jpg'
            bg_img = np_img(path_img)

        # 八、词云部分
        print('===================8、开始生成词云图====================')
        try:
            path_font = r'C:\Windows\Fonts\simhei.ttf'
            my_wordcloud(path_font, bg_img, keywords)
            print('程序执行完毕！')
        except Exception as e:
            print(e)
            print('词云图片生成失败，请检查错误！')
            err_time10 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
            with open(r'E:\练习文件\项目文件\python_log\cloud_analysis.txt', 'a', encoding='utf-8') as f_cloud2:
                f_cloud2.writelines('词云生成报错：%s  %s\n' % (e, err_time10))

    else:
        print('准备退出程序！')
        time.sleep(1.5)
        exit()